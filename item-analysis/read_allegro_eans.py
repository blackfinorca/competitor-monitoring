"""Read Allegro Excel export and write unique EANs to CSV.

Usage:
    python jobs/read_allegro_eans.py
    python jobs/read_allegro_eans.py --input path/to/file.xlsx --output data/eans.csv
"""

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

_DEFAULT_XLSX = "item-analysis/Allegro zalistované položky 42026.xlsx"
_DEFAULT_CSV = "item-analysis/allegro_eans.csv"


def main(xlsx_path: str = _DEFAULT_XLSX, csv_path: str = _DEFAULT_CSV) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        ean_col = headers.index("products_ean")
        title_col = headers.index("title")
        price_col = headers.index("price_sk")
    except ValueError as e:
        print(f"ERROR: missing column: {e}", file=sys.stderr)
        return 1

    seen: set[str] = set()
    rows: list[dict] = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_ean = row[ean_col]
        if raw_ean is None:
            skipped += 1
            continue
        ean = str(raw_ean).strip().split(".")[0]  # remove .0 from numeric EANs
        if not ean or ean in seen:
            continue
        seen.add(ean)
        rows.append({
            "ean": ean,
            "title": row[title_col] or "",
            "our_price_sk": row[price_col] or "",
        })

    wb.close()

    Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["ean", "title", "our_price_sk"])
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} unique EANs to {csv_path}  (skipped {skipped} rows without EAN)")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=_DEFAULT_XLSX)
    parser.add_argument("--output", default=_DEFAULT_CSV)
    args = parser.parse_args()
    sys.exit(main(args.input, args.output))
