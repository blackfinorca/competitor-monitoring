"""Export allegro_offers CSV to a wide-format Excel: one row per EAN, sellers as column groups.

Usage:
    python jobs/export_allegro_offers.py
    python jobs/export_allegro_offers.py --input data/allegro_offers.csv
    python jobs/export_allegro_offers.py --input data/allegro_offers.csv --output reports/allegro.xlsx

    # With reference prices from a CSV (ean + price column):
    python jobs/export_allegro_offers.py --reference data/allegro_eans.csv --ref-price-col our_price_sk --ref-label "Our Price"

    # With reference prices from an Excel sheet:
    python jobs/export_allegro_offers.py --reference path/to/file.xlsx --ref-sheet KUTILOVO --ref-ean-col EAN --ref-price-col Price
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_C_HEADER_BG  = "1F3864"
_C_HEADER_FG  = "FFFFFF"
_C_FIXED_BG   = "D6E4F0"
_C_FIXED_HDR  = "2E75B6"
_C_SELLER_HDR = "404040"
_C_SELLER_ALT = "F5F5F5"
_C_CHEAPEST   = "C6EFCE"
_C_CHEAPEST_FG = "276221"
_C_PRICIEST   = "FCE4D6"
_C_PRICIEST_FG = "843C0C"
_C_MISSING    = "EFEFEF"
_C_MISSING_FG = "AAAAAA"
_C_SUBHDR_BG  = "BDD7EE"
_C_REF_HDR    = "7B5EA7"   # purple — reference price column header
_C_REF_BG     = "EDE7F6"   # light purple — reference price cells

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=10, name="Calibri")


def _load_csv(path: Path) -> tuple[list[str], dict[str, dict[str, dict]], dict[str, float | None]]:
    """Return (ordered_eans, data, box_prices).

    data       = {ean: {seller: {title, price_eur, delivery_eur}}}
    box_prices = {ean: box_price_eur}  — one value per EAN
    Keeps first title seen per EAN; keeps lowest price when a seller appears twice.
    """
    data: dict[str, dict[str, dict]] = {}
    box_prices: dict[str, float | None] = {}
    ean_order: list[str] = []

    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ean = row["ean"].strip()
            seller = row["seller"].strip()
            if not ean or not seller:
                continue

            price = float(row["price_eur"]) if row.get("price_eur") else None
            delivery = float(row["delivery_eur"]) if row.get("delivery_eur") else None
            title = row.get("title", "").strip()
            raw_box = row.get("box_price_eur", "")
            try:
                box_price: float | None = float(raw_box) if raw_box else None
            except ValueError:
                box_price = None

            if ean not in data:
                data[ean] = {}
                ean_order.append(ean)

            # Store first non-None box price seen for this EAN
            if ean not in box_prices or (box_price is not None and box_prices[ean] is None):
                box_prices[ean] = box_price

            existing = data[ean].get(seller)
            total = (price or 0) + (delivery or 0)
            if existing is None or (price and existing["price"] and price < existing["price"]):
                data[ean][seller] = {
                    "title": title or (existing["title"] if existing else ""),
                    "price": price,
                    "delivery": delivery,
                    "total": total if price else None,
                }

    return ean_order, data, box_prices


def _title_for_ean(ean: str, data: dict[str, dict[str, dict]]) -> str:
    for seller_data in data[ean].values():
        if seller_data["title"]:
            return seller_data["title"]
    return ""


def load_reference_prices(
    ref_path: str,
    sheet: str | None = None,
    ean_col: str = "ean",
    price_col: str = "our_price_sk",
) -> dict[str, float | None]:
    """Load EAN → price mapping from a CSV or Excel reference file.

    Args:
        ref_path:  Path to .csv or .xlsx reference file.
        sheet:     Sheet name (Excel only; ignored for CSV).
        ean_col:   Column name containing the EAN.
        price_col: Column name containing the price.

    Returns:
        dict mapping EAN string → float price (or None if blank/unparseable).
    """
    path = Path(ref_path)
    prices: dict[str, float | None] = {}

    if path.suffix.lower() in (".xlsx", ".xls", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            ean_idx = headers.index(ean_col)
            price_idx = headers.index(price_col)
        except ValueError as e:
            raise ValueError(f"Column not found in '{sheet or ws.title}': {e}. Available: {headers}") from e
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_ean = row[ean_idx]
            if raw_ean is None:
                continue
            ean = str(raw_ean).strip().split(".")[0]
            if not ean:
                continue
            raw_price = row[price_idx]
            try:
                prices[ean] = float(raw_price) if raw_price is not None else None
            except (ValueError, TypeError):
                prices[ean] = None
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                ean = row.get(ean_col, "").strip().split(".")[0]
                if not ean:
                    continue
                raw_price = row.get(price_col, "")
                try:
                    prices[ean] = float(raw_price) if raw_price else None
                except (ValueError, TypeError):
                    prices[ean] = None

    matched = sum(1 for v in prices.values() if v is not None)
    print(f"  Reference '{path.name}': {len(prices)} EANs loaded, {matched} with prices")
    return prices


def build_workbook(
    ean_order: list[str],
    data: dict[str, dict[str, dict]],
    sellers: list[str],
    source_name: str,
    ref_prices: dict[str, float | None] | None = None,
    ref_label: str = "Reference",
    box_prices: dict[str, float | None] | None = None,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allegro Offers"

    HAS_BOX = box_prices is not None and any(v is not None for v in box_prices.values())
    HAS_REF = ref_prices is not None
    # Column layout: EAN | Title | [Box Price] | [Ref Price] | sellers…
    N_FIXED = 2 + (1 if HAS_BOX else 0) + (1 if HAS_REF else 0)
    N_PER_SELLER = 3                 # Price | Delivery | Total
    total_cols = N_FIXED + N_PER_SELLER * len(sellers)

    # ---- Row 1: title bar -----------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1)
    c.value = (
        f"Allegro.sk Competitor Prices  ·  {len(ean_order)} EANs  ·  "
        f"{len(sellers)} sellers  ·  {source_name}  ·  {date.today().isoformat()}"
    )
    c.font = Font(bold=True, color=_C_HEADER_FG, size=13, name="Calibri")
    c.fill = _fill(_C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ---- Row 2: seller group headers ------------------------------------
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_FIXED)
    lbl = ws.cell(row=2, column=1)
    lbl.value = "Product"
    lbl.font = Font(bold=True, color=_C_HEADER_FG, size=10, name="Calibri")
    lbl.fill = _fill(_C_FIXED_HDR)
    lbl.alignment = Alignment(horizontal="center", vertical="center")

    for i, seller in enumerate(sellers):
        col_start = N_FIXED + 1 + i * N_PER_SELLER
        col_end = col_start + N_PER_SELLER - 1
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        c = ws.cell(row=2, column=col_start)
        c.value = seller
        c.font = Font(bold=True, color=_C_HEADER_FG, size=10, name="Calibri")
        c.fill = _fill(_C_SELLER_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ---- Row 3: sub-headers --------------------------------------------
    fixed_headers = ["EAN", "Title"]
    if HAS_BOX:
        fixed_headers.append("Box Price (€)")
    if HAS_REF:
        fixed_headers.append(f"{ref_label} (€)")

    _C_BOX_HDR = "E67E22"   # orange — box price column
    _C_BOX_BG  = "FDEBD0"

    for col, hdr in enumerate(fixed_headers, start=1):
        is_box = HAS_BOX and hdr == "Box Price (€)"
        is_ref = HAS_REF and col == len(fixed_headers)
        c = ws.cell(row=3, column=col, value=hdr)
        c.font = Font(bold=True, size=9, name="Calibri",
                      color=_C_HEADER_FG if (is_box or is_ref) else "000000")
        c.fill = _fill(_C_BOX_HDR if is_box else (_C_REF_HDR if is_ref else _C_FIXED_BG))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER

    for i in range(len(sellers)):
        col_start = N_FIXED + 1 + i * N_PER_SELLER
        bg = _C_SUBHDR_BG if i % 2 == 0 else "D8D8D8"
        for j, sub in enumerate(["Price (€)", "Delivery (€)", "Total (€)"]):
            c = ws.cell(row=3, column=col_start + j, value=sub)
            c.font = Font(bold=True, size=9, name="Calibri")
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _BORDER
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}3"

    # ---- Data rows ------------------------------------------------------
    for row_idx, ean in enumerate(ean_order, start=4):
        row_bg = "FFFFFF" if row_idx % 2 == 0 else "F9FBFD"
        seller_data = data[ean]

        # Find cheapest and priciest total across sellers for this EAN
        totals = [v["total"] for v in seller_data.values() if v["total"] is not None]
        min_total = min(totals) if totals else None
        max_total = max(totals) if totals else None

        def _cell(col: int, value, fmt=None, bg=row_bg):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.font = Font(size=9, name="Calibri")
            c.fill = _fill(bg)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")
            if fmt:
                c.number_format = fmt
            return c

        _cell(1, ean).alignment = Alignment(vertical="center", horizontal="center")
        title_c = _cell(2, _title_for_ean(ean, data))
        title_c.alignment = Alignment(wrap_text=True, vertical="center")

        next_col = 3
        if HAS_BOX:
            box_val = box_prices.get(ean) if box_prices else None
            bc = _cell(next_col, box_val, fmt='#,##0.00 "€"', bg=_C_BOX_BG)
            bc.font = Font(bold=True, size=9, name="Calibri",
                           color="000000" if box_val is not None else _C_MISSING_FG)
            if box_val is None:
                bc.value = "—"
                bc.alignment = Alignment(horizontal="center", vertical="center")
            next_col += 1

        if HAS_REF:
            ref_val = ref_prices.get(ean) if ref_prices else None
            rc = _cell(next_col, ref_val, fmt='#,##0.00 "€"', bg=_C_REF_BG)
            rc.font = Font(bold=True, size=9, name="Calibri",
                           color="000000" if ref_val is not None else _C_MISSING_FG)
            if ref_val is None:
                rc.value = "—"
                rc.alignment = Alignment(horizontal="center", vertical="center")

        for i, seller in enumerate(sellers):
            col_start = N_FIXED + 1 + i * N_PER_SELLER
            comp_bg = _C_SELLER_ALT if i % 2 == 0 else "FAFAFA"
            sd = seller_data.get(seller)

            if sd is None:
                for j in range(N_PER_SELLER):
                    c = ws.cell(row=row_idx, column=col_start + j, value="—")
                    c.font = Font(color=_C_MISSING_FG, size=9, name="Calibri")
                    c.fill = _fill(_C_MISSING)
                    c.border = _BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # Price
            _cell(col_start, sd["price"], fmt='#,##0.00 "€"', bg=comp_bg)

            # Delivery
            _cell(col_start + 1, sd["delivery"], fmt='#,##0.00 "€"', bg=comp_bg)

            # Total — colour-coded
            total = sd["total"]
            if total is not None and min_total is not None and max_total is not None and min_total != max_total:
                if abs(total - min_total) < 0.005:
                    total_bg, total_fg = _C_CHEAPEST, _C_CHEAPEST_FG
                elif abs(total - max_total) < 0.005:
                    total_bg, total_fg = _C_PRICIEST, _C_PRICIEST_FG
                else:
                    total_bg, total_fg = comp_bg, "000000"
            else:
                total_bg, total_fg = comp_bg, "000000"

            tc = ws.cell(row=row_idx, column=col_start + 2, value=total)
            tc.number_format = '#,##0.00 "€"'
            tc.font = Font(bold=True, color=total_fg, size=9, name="Calibri")
            tc.fill = _fill(total_bg)
            tc.border = _BORDER
            tc.alignment = Alignment(vertical="center")

    # ---- Column widths --------------------------------------------------
    ws.column_dimensions["A"].width = 16   # EAN
    ws.column_dimensions["B"].width = 40   # Title
    fixed_col = 3
    if HAS_BOX:
        ws.column_dimensions[get_column_letter(fixed_col)].width = 13  # Box price
        fixed_col += 1
    if HAS_REF:
        ws.column_dimensions[get_column_letter(fixed_col)].width = 14  # Ref price
    for i in range(len(sellers)):
        col_start = N_FIXED + 1 + i * N_PER_SELLER
        ws.column_dimensions[get_column_letter(col_start)].width = 11      # Price
        ws.column_dimensions[get_column_letter(col_start + 1)].width = 12  # Delivery
        ws.column_dimensions[get_column_letter(col_start + 2)].width = 11  # Total

    return wb


def main(
    input_path: str,
    output_path: str | None = None,
    ref_path: str | None = None,
    ref_sheet: str | None = None,
    ref_ean_col: str = "ean",
    ref_price_col: str = "our_price_sk",
    ref_label: str = "Reference",
) -> None:
    path = Path(input_path)
    if not path.exists():
        print(f"ERROR: {input_path} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {input_path} …")
    ean_order, data, box_prices = _load_csv(path)
    sellers: list[str] = sorted({s for ean_data in data.values() for s in ean_data})
    box_count = sum(1 for v in box_prices.values() if v is not None)
    print(f"  {len(ean_order)} EANs  ·  {len(sellers)} sellers  ·  {box_count} box prices")

    ref_prices = None
    if ref_path:
        print(f"Loading reference prices from {ref_path} …")
        ref_prices = load_reference_prices(ref_path, ref_sheet, ref_ean_col, ref_price_col)

    if not output_path:
        stem = path.stem
        output_path = f"reports/allegro_{stem}_{date.today().isoformat()}.xlsx"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    print("Building workbook …")
    wb = build_workbook(ean_order, data, sellers, path.name, ref_prices, ref_label, box_prices)
    wb.save(output_path)
    print(f"Saved → {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="data/allegro_offers.csv")
    parser.add_argument("--output", default=None)
    parser.add_argument("--reference", dest="ref_path", default=None,
                        help="CSV or Excel file with reference prices (e.g. KUTILOVO price list)")
    parser.add_argument("--ref-sheet", default=None,
                        help="Sheet name in reference Excel (e.g. KUTILOVO)")
    parser.add_argument("--ref-ean-col", default="ean",
                        help="EAN column name in reference file (default: ean)")
    parser.add_argument("--ref-price-col", default="our_price_sk",
                        help="Price column name in reference file (default: our_price_sk)")
    parser.add_argument("--ref-label", default="Reference",
                        help="Column header label for reference price (default: Reference)")
    args = parser.parse_args()
    main(
        args.input, args.output,
        args.ref_path, args.ref_sheet,
        args.ref_ean_col, args.ref_price_col,
        args.ref_label,
    )
