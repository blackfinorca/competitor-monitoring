"""Backfill missing product EANs from Allegro exports.

Reads an Allegro Excel or CSV export containing at least EAN and title, finds
the best matching AG product by title, and writes the Allegro EAN back into
`products.ean` only when the product currently has no EAN.
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from agnaradie_pricing.catalogue.normalise import fold_diacritics, normalise_brand, normalise_ean
from agnaradie_pricing.db.models import Product
from agnaradie_pricing.db.session import make_engine
from agnaradie_pricing.matching.llm_matcher import OpenAIClient, find_best_llm_match
from agnaradie_pricing.matching.vector_search import TitleVectorIndex
from agnaradie_pricing.settings import Settings

logger = logging.getLogger(__name__)

_DEFAULT_INPUT = Path("item-analysis/Allegro zalistované položky 42026.xlsx")
_DEFAULT_REPORT_DIR = Path("reports")
_VECTOR_LIMIT = 40
_LLM_LIMIT = 40
_DETERMINISTIC_SCORE = 0.93
_DETERMINISTIC_MARGIN = 0.05
_LLM_MIN_CONFIDENCE = 0.81
_LLM_MIN_SCORE = 0.80
_MODEL_TOKEN_RE = re.compile(r"[a-z0-9][a-z0-9./-]*\d[a-z0-9./-]*", re.IGNORECASE)
_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")


def load_source_rows(path: str | Path) -> list[dict[str, str]]:
    input_path = Path(path)
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_rows(input_path)
    if suffix == ".csv":
        return _load_csv_rows(input_path)
    raise ValueError(f"Unsupported input format: {input_path}")


def _load_xlsx_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        ean_col = headers.index("products_ean")
        title_col = headers.index("title")
        price_col = headers.index("price_sk")
    except ValueError as exc:
        wb.close()
        raise ValueError(f"{path} missing expected Allegro columns") from exc

    rows: list[dict[str, str]] = []
    seen_pairs: set[tuple[str, str]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ean = normalise_ean(row[ean_col])
        title = str(row[title_col] or "").strip()
        if not ean or not title:
            continue
        pair = (ean, title)
        if pair in seen_pairs:
            continue
        seen_pairs.add(pair)
        rows.append(
            {
                "ean": ean,
                "title": title,
                "allegro_price_sk": _stringify_value(row[price_col]),
            }
        )
    wb.close()
    return rows


def _load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows: list[dict[str, str]] = []
        seen_pairs: set[tuple[str, str]] = set()
        for row in reader:
            ean = normalise_ean(row.get("ean"))
            title = str(row.get("title") or "").strip()
            if not ean or not title:
                continue
            pair = (ean, title)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            rows.append(
                {
                    "ean": ean,
                    "title": title,
                    "allegro_price_sk": _stringify_value(row.get("our_price_sk") or row.get("price_sk")),
                }
            )
    return rows


def _stringify_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _iter_source_rows(rows: list[dict[str, str]], limit: int | None = None):
    count = 0
    for row in rows:
        if limit is not None and count >= limit:
            break
        yield row
        count += 1


def _normalise_title(title: str) -> str:
    folded = fold_diacritics(title.lower())
    collapsed = _NON_ALNUM_RE.sub(" ", folded)
    return " ".join(collapsed.split())


def _extract_model_tokens(title: str) -> set[str]:
    folded = fold_diacritics(title.lower())
    return {_dense_token(token) for token in _MODEL_TOKEN_RE.findall(folded) if _dense_token(token)}


def _dense_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", fold_diacritics(value.lower()))


def _candidate_text(product: dict[str, Any]) -> str:
    return _normalise_title(f"{product.get('title') or ''} {product.get('mpn') or ''}")


def _tokens_present(title: str, product: dict[str, Any]) -> bool:
    model_tokens = _extract_model_tokens(title)
    if not model_tokens:
        return True
    searchable = _dense_token(f"{product.get('title') or ''} {product.get('mpn') or ''}")
    return all(token in searchable for token in model_tokens)


def _detect_brand(row: dict[str, str], known_brands: dict[str, str]) -> str | None:
    tokens = set(_normalise_title(row["title"]).split())
    matched = {brand for token, brand in known_brands.items() if token in tokens}
    if len(matched) == 1:
        return next(iter(matched))
    return None


def _build_brand_token_map(products: list[dict[str, Any]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for product in products:
        brand = normalise_brand(product.get("brand"))
        if not brand:
            continue
        for token in _normalise_title(brand).split():
            if len(token) < 3:
                continue
            mapping.setdefault(token, brand)
    return mapping


def _load_target_products(session: Session) -> list[dict[str, Any]]:
    rows = session.execute(
        select(
            Product.id,
            Product.sku,
            Product.brand,
            Product.mpn,
            Product.ean,
            Product.title,
        ).where(Product.ean.is_(None), Product.title.is_not(None))
    ).fetchall()
    return [dict(row._mapping) for row in rows]


def _build_exact_title_index(products: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        title = str(product.get("title") or "").strip()
        if title:
            index[_normalise_title(title)].append(product)
    return index


def _restrict_by_brand(
    brand: str | None,
    scored_candidates: list[tuple[dict[str, Any], float]],
) -> list[tuple[dict[str, Any], float]]:
    if not brand:
        return scored_candidates
    filtered = [
        (candidate, score)
        for candidate, score in scored_candidates
        if normalise_brand(candidate.get("brand")) == brand
    ]
    return filtered or scored_candidates


def _find_exact_title_match(
    row: dict[str, str],
    title_index: dict[str, list[dict[str, Any]]],
    detected_brand: str | None,
) -> dict[str, Any] | None:
    matches = title_index.get(_normalise_title(row["title"]), [])
    if detected_brand:
        matches = [match for match in matches if normalise_brand(match.get("brand")) == detected_brand]
    if len(matches) == 1:
        return matches[0]
    return None


def _vector_candidates(
    row: dict[str, str],
    index: TitleVectorIndex,
    *,
    limit: int,
) -> list[tuple[dict[str, Any], float]]:
    if hasattr(index, "search_with_scores"):
        return list(index.search_with_scores({"title": row["title"]}, limit=limit))
    candidates = index.search({"title": row["title"]}, limit=limit)
    return [(candidate, 0.0) for candidate in candidates]


def _deterministic_match(
    row: dict[str, str],
    scored_candidates: list[tuple[dict[str, Any], float]],
) -> tuple[dict[str, Any], str, float] | None:
    if not scored_candidates:
        return None
    top_candidate, top_score = scored_candidates[0]
    second_score = scored_candidates[1][1] if len(scored_candidates) > 1 else 0.0
    if (
        top_score >= _DETERMINISTIC_SCORE
        and (top_score - second_score) >= _DETERMINISTIC_MARGIN
        and _tokens_present(row["title"], top_candidate)
    ):
        return top_candidate, "vector_strict", top_score
    return None


def _llm_match(
    row: dict[str, str],
    scored_candidates: list[tuple[dict[str, Any], float]],
    *,
    llm_client,
) -> tuple[dict[str, Any], str, float] | None:
    if not scored_candidates:
        return None
    llm_candidates = [candidate for candidate, _score in scored_candidates[:_LLM_LIMIT]]
    match = find_best_llm_match({"title": row["title"]}, llm_candidates, llm_client=llm_client)
    if match is None:
        return None
    candidate, (match_type, confidence) = match
    score_map = {product["id"]: score for product, score in scored_candidates}
    candidate_score = score_map.get(candidate["id"], 0.0)
    if (
        confidence >= _LLM_MIN_CONFIDENCE
        and candidate_score >= _LLM_MIN_SCORE
        and _tokens_present(row["title"], candidate)
    ):
        return candidate, match_type, confidence
    return None


def _match_row(
    row: dict[str, str],
    *,
    index: TitleVectorIndex,
    title_index: dict[str, list[dict[str, Any]]],
    known_brands: dict[str, str],
    llm_client=None,
) -> tuple[str, dict[str, Any] | None, float, str]:
    detected_brand = _detect_brand(row, known_brands)
    exact_match = _find_exact_title_match(row, title_index, detected_brand)
    if exact_match is not None:
        return "matched", exact_match, 1.0, "exact_title"

    scored_candidates = _restrict_by_brand(
        detected_brand,
        _vector_candidates(row, index, limit=_VECTOR_LIMIT),
    )
    deterministic = _deterministic_match(row, scored_candidates)
    if deterministic is not None:
        product, match_type, confidence = deterministic
        return "matched", product, confidence, match_type

    if llm_client is not None:
        llm_match = _llm_match(row, scored_candidates, llm_client=llm_client)
        if llm_match is not None:
            product, match_type, confidence = llm_match
            return "matched", product, confidence, match_type

    if scored_candidates:
        return "ambiguous", None, 0.0, ""
    return "no_match", None, 0.0, ""


def _report_row(
    row: dict[str, str],
    *,
    status: str,
    product: dict[str, Any] | None = None,
    confidence: float = 0.0,
    match_type: str = "",
) -> dict[str, Any]:
    return {
        "status": status,
        "product_id": product.get("id") if product else "",
        "sku": product.get("sku") if product else "",
        "old_ean": product.get("ean") if product and product.get("ean") else "",
        "new_ean": row["ean"],
        "confidence": f"{confidence:.2f}" if confidence else "",
        "match_type": match_type,
        "allegro_title": row["title"],
        "product_title": product.get("title") if product else "",
        "allegro_price_sk": row.get("allegro_price_sk", ""),
    }


def _flush_updates(session: Session, updates: list[dict[str, Any]], *, dry_run: bool) -> int:
    if not updates:
        return 0
    if dry_run:
        return len(updates)
    applied = 0
    for update in updates:
        result = session.execute(
            text("UPDATE products SET ean=:ean WHERE id=:id AND ean IS NULL"),
            update,
        )
        applied += result.rowcount or 0
    session.commit()
    return applied


def _default_report_path() -> Path:
    return _DEFAULT_REPORT_DIR / f"allegro_ean_backfill_{date.today().isoformat()}.csv"


def _write_report(rows: list[dict[str, Any]], path: str | Path) -> None:
    report_path = Path(path)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "status",
                "product_id",
                "sku",
                "old_ean",
                "new_ean",
                "confidence",
                "match_type",
                "allegro_title",
                "product_title",
                "allegro_price_sk",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def main(
    input_path: str | Path = _DEFAULT_INPUT,
    *,
    dry_run: bool = False,
    limit: int | None = None,
    batch_size: int = 100,
    llm: bool = False,
    report_path: str | Path | None = None,
) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    settings = Settings()
    report_output = Path(report_path) if report_path is not None else _default_report_path()
    source_rows = load_source_rows(input_path)
    remaining_by_ean = Counter(row["ean"] for row in source_rows[:limit] if row.get("ean")) if limit else Counter(
        row["ean"] for row in source_rows if row.get("ean")
    )

    engine = make_engine(settings)
    llm_client = None
    if llm:
        llm_client = OpenAIClient(
            api_key=getattr(settings, "openai_api_key", None) or "missing",
            model=getattr(settings, "openai_model", "gpt-5-nano"),
        )

    report_rows: list[dict[str, Any]] = []
    buffered_updates: list[dict[str, Any]] = []
    product_claims: dict[int, str] = {}
    applied = 0
    pending_ean_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)

    with Session(engine) as session:
        products = _load_target_products(session)
        title_index = _build_exact_title_index(products)
        known_brands = _build_brand_token_map(products)
        index = TitleVectorIndex(products)

        try:
            for row in _iter_source_rows(source_rows, limit=limit):
                status, product, confidence, match_type = _match_row(
                    row,
                    index=index,
                    title_index=title_index,
                    known_brands=known_brands,
                    llm_client=llm_client,
                )
                report = _report_row(
                    row,
                    status=status,
                    product=product,
                    confidence=confidence,
                    match_type=match_type,
                )
                pending_ean_rows[row["ean"]].append(report)
                if status == "matched" and product is not None:
                    claimed_ean = product_claims.get(product["id"])
                    if claimed_ean is not None and claimed_ean != row["ean"]:
                        report["status"] = "product_conflict"

                remaining_by_ean[row["ean"]] -= 1
                if remaining_by_ean[row["ean"]] > 0:
                    continue

                ean = row["ean"]
                ean_reports = pending_ean_rows.pop(ean, [])
                matched_product_ids = {
                    int(report["product_id"])
                    for report in ean_reports
                    if str(report.get("status")) == "matched" and report.get("product_id")
                }
                if len(matched_product_ids) > 1:
                    for ean_report in ean_reports:
                        ean_report["status"] = "duplicate_ean_conflict"
                        ean_report["match_type"] = ""
                        ean_report["confidence"] = ""
                    report_rows.extend(ean_reports)
                    continue

                if len(matched_product_ids) == 1:
                    product_id = next(iter(matched_product_ids))
                    conflicting_claim = product_claims.get(product_id)
                    if conflicting_claim is not None and conflicting_claim != ean:
                        for ean_report in ean_reports:
                            ean_report["status"] = "product_conflict"
                            ean_report["match_type"] = ""
                            ean_report["confidence"] = ""
                    else:
                        product_claims[product_id] = ean
                        buffered_updates.append({"id": product_id, "ean": ean})
                        report_rows.extend(ean_reports)
                        if len(buffered_updates) >= max(batch_size, 1):
                            applied += _flush_updates(session, buffered_updates, dry_run=dry_run)
                            buffered_updates.clear()
                        continue

                report_rows.extend(ean_reports)

        except KeyboardInterrupt:
            applied += _flush_updates(session, buffered_updates, dry_run=dry_run)
            buffered_updates.clear()
            for unfinished_reports in pending_ean_rows.values():
                report_rows.extend(unfinished_reports)
            _write_report(report_rows, report_output)
            logger.warning("Interrupted after %d applied updates", applied)
            return 130

        applied += _flush_updates(session, buffered_updates, dry_run=dry_run)
        buffered_updates.clear()

    _write_report(report_rows, report_output)
    logger.info("Processed %d Allegro rows; applied %d updates", len(report_rows), applied)
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=str(_DEFAULT_INPUT))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int)
    parser.add_argument("--batch-size", type=int, default=100)
    parser.add_argument("--llm", action="store_true")
    parser.add_argument("--report")
    args = parser.parse_args()
    sys.exit(
        main(
            input_path=args.input,
            dry_run=args.dry_run,
            limit=args.limit,
            batch_size=args.batch_size,
            llm=args.llm,
            report_path=args.report,
        )
    )
