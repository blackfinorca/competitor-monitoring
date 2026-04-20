"""Export manufacturer price comparison to a formatted Excel workbook.

One row per ToolZone product; one column group per competitor.
Colour-codes price differences: green = we are cheaper, red = we are more expensive.

Usage
-----
    python jobs/export_manufacturer.py --manufacturer knipex
    python jobs/export_manufacturer.py --manufacturer knipex --only boukal_cz bo_import_cz
    python jobs/export_manufacturer.py --manufacturer knipex --min-confidence 0.85 --output reports/knipex.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from agnaradie_pricing.db.session import make_session_factory
from agnaradie_pricing.settings import Settings, load_competitors, own_store_ids

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

_C_HEADER_BG   = "1F3864"   # dark navy — header row
_C_HEADER_FG   = "FFFFFF"   # white text
_C_TZ_BG       = "D6E4F0"   # light blue — ToolZone columns
_C_TZ_HDR      = "2E75B6"   # medium blue — ToolZone header
_C_COMP_HDR    = "404040"   # dark grey — competitor header
_C_COMP_ALT    = "F5F5F5"   # very light grey — alternating competitor groups
_C_CHEAPER     = "C6EFCE"   # green fill — competitor cheaper than TZ
_C_CHEAPER_FG  = "276221"
_C_EXPENSIVE   = "FCE4D6"   # orange fill — competitor more expensive than TZ
_C_EXPENSIVE_FG = "843C0C"
_C_EQUAL       = "FFF2CC"   # yellow fill — within 1%
_C_EQUAL_FG    = "7D6608"
_C_NO_MATCH    = "EFEFEF"   # light grey — no match
_C_OOS         = "BFBFBF"   # grey text — out of stock
_C_TITLE_BG    = "1F3864"
_C_SUBHDR_BG   = "BDD7EE"   # light blue — sub-header row

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_data(
    manufacturer: str,
    competitor_ids: list[str],
    min_confidence: float,
    session,
) -> tuple[list[dict], dict[int, dict[str, dict]]]:
    """Return (tz_products, matches).

    tz_products : list of dicts — one per ToolZone listing, sorted by title
    matches     : {tz_id: {competitor_id: {price, stock, url, match_type, confidence}}}
    """
    brand_like = f"%{manufacturer}%"

    tz_rows = session.execute(
        text("""
            SELECT id, ean, mpn, title, price_eur, in_stock, url, scraped_at
            FROM competitor_listings
            WHERE competitor_id = 'toolzone_sk'
              AND LOWER(brand) LIKE :brand
            ORDER BY title
        """),
        {"brand": brand_like},
    ).fetchall()

    if not tz_rows:
        return [], {}

    tz_ids = [r.id for r in tz_rows]
    tz_params: dict = {f"tz_{i}": v for i, v in enumerate(tz_ids)}
    tz_placeholders = ", ".join(f":tz_{i}" for i in range(len(tz_ids)))

    comp_filter = ""
    comp_params: dict = {}
    if competitor_ids:
        cid_params = {f"cid_{i}": v for i, v in enumerate(competitor_ids)}
        cid_placeholders = ", ".join(f":cid_{i}" for i in range(len(competitor_ids)))
        comp_filter = f"AND cl.competitor_id IN ({cid_placeholders})"
        comp_params = cid_params

    match_rows = session.execute(
        text(f"""
            SELECT
                lm.toolzone_listing_id  AS tz_id,
                cl.competitor_id,
                cl.price_eur            AS comp_price,
                cl.in_stock             AS comp_stock,
                cl.url                  AS comp_url,
                cl.scraped_at           AS comp_scraped,
                lm.match_type,
                lm.confidence
            FROM listing_matches lm
            JOIN competitor_listings cl ON cl.id = lm.competitor_listing_id
            WHERE lm.toolzone_listing_id IN ({tz_placeholders})
              AND lm.confidence >= :min_conf
              {comp_filter}
        """),
        {"min_conf": min_confidence, **tz_params, **comp_params},
    ).fetchall()

    tz_products = [
        {
            "id":         r.id,
            "ean":        r.ean or "",
            "mpn":        r.mpn or "",
            "title":      r.title or "",
            "price_eur":  float(r.price_eur) if r.price_eur is not None else None,
            "in_stock":   r.in_stock,
            "url":        r.url or "",
            "scraped_at": r.scraped_at,
        }
        for r in tz_rows
    ]

    matches: dict[int, dict[str, dict]] = {}
    for row in match_rows:
        cid = row.competitor_id
        tz_id = row.tz_id
        if tz_id not in matches:
            matches[tz_id] = {}
        existing = matches[tz_id].get(cid)
        price = float(row.comp_price) if row.comp_price is not None else None
        # Keep lowest price when same competitor appears via multiple matches
        if existing is None or (
            price is not None
            and (existing["price"] is None or price < existing["price"])
        ):
            matches[tz_id][cid] = {
                "price":      price,
                "stock":      row.comp_stock,
                "url":        row.comp_url or "",
                "match_type": row.match_type,
                "confidence": float(row.confidence),
            }

    return tz_products, matches


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

_MATCH_BADGE = {
    "exact_ean":          "EAN ✓",
    "exact_mpn":          "MPN ✓",
    "mpn_no_brand":       "MPN ~",
    "regex_ean_title":    "Regex ~",
    "regex_mpn_title":    "Regex ~",
    "regex_mpn_no_brand": "Regex ~",
    "llm_fuzzy":          "LLM ~",
}


def _diff_pct(tz_price: float | None, comp_price: float | None) -> float | None:
    if tz_price and comp_price is not None:
        return (comp_price - tz_price) / tz_price * 100
    return None


def build_workbook(
    manufacturer: str,
    tz_products: list[dict],
    matches: dict[int, dict[str, dict]],
    competitor_ids: list[str],
    competitor_names: dict[str, str],
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{manufacturer.title()} prices"

    # ---- Column layout -------------------------------------------------
    # Fixed columns: Title | EAN | MPN | TZ Price | TZ Stock | TZ URL
    # Per competitor: Price | Diff% | Match | Stock | URL
    N_FIXED = 6
    N_PER_COMP = 5
    FIXED_HEADERS = ["Title", "EAN", "MPN", "TZ Price (€)", "TZ In Stock", "TZ URL"]
    COMP_SUB = ["Price (€)", "vs TZ", "Match", "In Stock", "URL"]

    total_cols = N_FIXED + N_PER_COMP * len(competitor_ids)

    # ---- Row 1: main title ---------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (
        f"{manufacturer.upper()}  —  Price Comparison  ·  "
        f"{len(tz_products)} ToolZone products  ·  {date.today().isoformat()}"
    )
    title_cell.font = Font(bold=True, color=_C_HEADER_FG, size=13, name="Calibri")
    title_cell.fill = _fill(_C_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ---- Row 2: competitor group headers --------------------------------
    # Fixed section label
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_FIXED)
    tz_lbl = ws.cell(row=2, column=1)
    tz_lbl.value = "ToolZone (reference)"
    tz_lbl.font = Font(bold=True, color=_C_HEADER_FG, size=10, name="Calibri")
    tz_lbl.fill = _fill(_C_TZ_HDR)
    tz_lbl.alignment = Alignment(horizontal="center", vertical="center")

    for i, cid in enumerate(competitor_ids):
        col_start = N_FIXED + 1 + i * N_PER_COMP
        col_end = col_start + N_PER_COMP - 1
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        cell = ws.cell(row=2, column=col_start)
        cell.value = competitor_names.get(cid, cid)
        cell.font = Font(bold=True, color=_C_HEADER_FG, size=10, name="Calibri")
        cell.fill = _fill(_C_COMP_HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ---- Row 3: column sub-headers --------------------------------------
    for col, hdr in enumerate(FIXED_HEADERS, start=1):
        cell = ws.cell(row=3, column=col, value=hdr)
        cell.font = Font(bold=True, color="000000", size=9, name="Calibri")
        cell.fill = _fill(_C_TZ_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for i, cid in enumerate(competitor_ids):
        col_start = N_FIXED + 1 + i * N_PER_COMP
        bg = _C_SUBHDR_BG if i % 2 == 0 else "D8D8D8"
        for j, sub in enumerate(COMP_SUB):
            cell = ws.cell(row=3, column=col_start + j, value=sub)
            cell.font = Font(bold=True, color="000000", size=9, name="Calibri")
            cell.fill = _fill(bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
    ws.row_dimensions[3].height = 28

    # ---- Freeze panes above data rows ------------------------------------
    ws.freeze_panes = "A4"

    # ---- Auto filter on row 3 -------------------------------------------
    ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}3"

    # ---- Data rows -------------------------------------------------------
    for row_idx, product in enumerate(tz_products, start=4):
        tz_price = product["price_eur"]
        tz_id = product["id"]
        row_matches = matches.get(tz_id, {})

        # Alternating row background
        row_bg = "FFFFFF" if row_idx % 2 == 0 else "F9FBFD"

        def _cell(col: int, value, number_format=None, hyperlink=None):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.font = _font(size=9)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")
            if number_format:
                c.number_format = number_format
            if hyperlink:
                c.hyperlink = hyperlink
                c.font = Font(color="0563C1", underline="single", size=9, name="Calibri")
            return c

        # Fixed columns
        title_c = _cell(1, product["title"])
        title_c.fill = _fill(row_bg)
        title_c.alignment = Alignment(wrap_text=True, vertical="center")

        _cell(2, product["ean"]).fill = _fill(row_bg)
        _cell(3, product["mpn"]).fill = _fill(row_bg)

        price_c = _cell(4, tz_price, number_format='#,##0.00 "€"')
        price_c.fill = _fill(_C_TZ_BG)
        price_c.font = Font(bold=True, size=9, name="Calibri")

        stock_val = "✅" if product["in_stock"] else ("❌" if product["in_stock"] is False else "—")
        stock_c = _cell(5, stock_val)
        stock_c.fill = _fill(_C_TZ_BG)
        stock_c.alignment = Alignment(horizontal="center", vertical="center")

        url_c = _cell(6, "Open ↗" if product["url"] else "—", hyperlink=product["url"] or None)
        url_c.fill = _fill(_C_TZ_BG)

        # Competitor columns
        for i, cid in enumerate(competitor_ids):
            col_start = N_FIXED + 1 + i * N_PER_COMP
            comp_bg = _C_COMP_ALT if i % 2 == 0 else "FAFAFA"

            match = row_matches.get(cid)
            if match is None:
                # No match — grey out the group
                for j in range(N_PER_COMP):
                    c = ws.cell(row=row_idx, column=col_start + j, value="—")
                    c.font = Font(color=_C_OOS, size=9, name="Calibri")
                    c.fill = _fill(_C_NO_MATCH)
                    c.border = _BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center")
                continue

            comp_price = match["price"]
            diff = _diff_pct(tz_price, comp_price)

            # Diff colour
            if diff is None:
                diff_fill, diff_fg = _C_NO_MATCH, "888888"
            elif diff < -1.0:
                diff_fill, diff_fg = _C_CHEAPER, _C_CHEAPER_FG    # competitor cheaper
            elif diff > 1.0:
                diff_fill, diff_fg = _C_EXPENSIVE, _C_EXPENSIVE_FG  # competitor expensive
            else:
                diff_fill, diff_fg = _C_EQUAL, _C_EQUAL_FG           # within 1%

            # Price cell
            p_cell = ws.cell(row=row_idx, column=col_start, value=comp_price)
            p_cell.number_format = '#,##0.00 "€"'
            p_cell.font = Font(bold=True, size=9, name="Calibri",
                               color=_C_OOS if not match["stock"] else "000000")
            p_cell.fill = _fill(comp_bg)
            p_cell.border = _BORDER
            p_cell.alignment = Alignment(vertical="center")

            # Diff% cell
            diff_str = f"{diff:+.1f}%" if diff is not None else "—"
            d_cell = ws.cell(row=row_idx, column=col_start + 1, value=diff_str)
            d_cell.font = Font(bold=True, color=diff_fg, size=9, name="Calibri")
            d_cell.fill = _fill(diff_fill)
            d_cell.border = _BORDER
            d_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Match type
            badge = _MATCH_BADGE.get(match["match_type"], match["match_type"])
            m_cell = ws.cell(row=row_idx, column=col_start + 2, value=badge)
            m_cell.font = Font(color="444444", size=8, name="Calibri")
            m_cell.fill = _fill(comp_bg)
            m_cell.border = _BORDER
            m_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Stock
            s_val = "✅" if match["stock"] else ("❌" if match["stock"] is False else "—")
            s_cell = ws.cell(row=row_idx, column=col_start + 3, value=s_val)
            s_cell.fill = _fill(comp_bg)
            s_cell.border = _BORDER
            s_cell.alignment = Alignment(horizontal="center", vertical="center")

            # URL
            u_cell = ws.cell(row=row_idx, column=col_start + 4,
                             value="Open ↗" if match["url"] else "—")
            if match["url"]:
                u_cell.hyperlink = match["url"]
                u_cell.font = Font(color="0563C1", underline="single", size=9, name="Calibri")
            else:
                u_cell.font = Font(color="888888", size=9, name="Calibri")
            u_cell.fill = _fill(comp_bg)
            u_cell.border = _BORDER
            u_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Column widths --------------------------------------------------
    ws.column_dimensions["A"].width = 42   # Title
    ws.column_dimensions["B"].width = 16   # EAN
    ws.column_dimensions["C"].width = 16   # MPN
    ws.column_dimensions["D"].width = 13   # TZ Price
    ws.column_dimensions["E"].width = 10   # TZ Stock
    ws.column_dimensions["F"].width = 10   # TZ URL

    for i in range(len(competitor_ids)):
        col_start = N_FIXED + 1 + i * N_PER_COMP
        ws.column_dimensions[get_column_letter(col_start)].width = 12      # Price
        ws.column_dimensions[get_column_letter(col_start + 1)].width = 9   # Diff%
        ws.column_dimensions[get_column_letter(col_start + 2)].width = 9   # Match
        ws.column_dimensions[get_column_letter(col_start + 3)].width = 8   # Stock
        ws.column_dimensions[get_column_letter(col_start + 4)].width = 9   # URL

    # ---- Summary sheet --------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    _build_summary_sheet(ws2, manufacturer, tz_products, matches, competitor_ids, competitor_names)

    return wb


def _build_summary_sheet(
    ws,
    manufacturer: str,
    tz_products: list[dict],
    matches: dict[int, dict[str, dict]],
    competitor_ids: list[str],
    competitor_names: dict[str, str],
) -> None:
    """Per-competitor summary: total matched, avg price diff, cheaper/equal/more expensive."""
    ws.title = "Summary"
    headers = ["Competitor", "Matched", "Match %", "Avg Price Diff", "Cheaper", "Equal (±1%)", "More Expensive", "Avg Conf."]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color=_C_HEADER_FG, size=10, name="Calibri")
        c.fill = _fill(_C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
    ws.row_dimensions[1].height = 22

    total_tz = len(tz_products)

    for row_idx, cid in enumerate(competitor_ids, start=2):
        matched = []
        for product in tz_products:
            m = matches.get(product["id"], {}).get(cid)
            if m:
                matched.append(m)

        n = len(matched)
        match_pct = n / total_tz * 100 if total_tz else 0

        diffs = [_diff_pct(p["price_eur"], m["price"])
                 for p, m in zip(tz_products, [matches.get(p["id"], {}).get(cid) for p in tz_products])
                 if m and p["price_eur"] and m["price"]]

        avg_diff = sum(diffs) / len(diffs) if diffs else None
        cheaper  = sum(1 for d in diffs if d < -1.0)
        equal    = sum(1 for d in diffs if -1.0 <= d <= 1.0)
        more_exp = sum(1 for d in diffs if d > 1.0)
        avg_conf = (sum(m["confidence"] for m in matched) / n) if n else None

        row_bg = "FFFFFF" if row_idx % 2 == 0 else _C_COMP_ALT

        def _sc(col, value, fmt=None):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.font = _font(size=10)
            c.fill = _fill(row_bg)
            c.border = _BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fmt:
                c.number_format = fmt
            return c

        _sc(1, competitor_names.get(cid, cid)).alignment = Alignment(vertical="center")
        _sc(2, n)
        _sc(3, round(match_pct, 1), fmt='0.0"%"')
        diff_c = _sc(4, round(avg_diff, 1) if avg_diff is not None else None, fmt='+0.0;-0.0;0.0')
        if avg_diff is not None:
            diff_c.fill = _fill(_C_CHEAPER if avg_diff < -1 else (_C_EXPENSIVE if avg_diff > 1 else _C_EQUAL))
        _sc(5, cheaper)
        _sc(6, equal)
        _sc(7, more_exp)
        _sc(8, round(avg_conf, 2) if avg_conf else None, fmt='0.00')

    col_widths = [22, 10, 10, 16, 10, 14, 16, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(
    manufacturer: str,
    output: Path | None = None,
    only: list[str] | None = None,
    min_confidence: float = 0.72,
) -> Path:
    settings = Settings()
    factory = make_session_factory(settings)
    all_competitors = load_competitors()
    own = own_store_ids()

    # Build ordered competitor list (exclude own stores and toolzone reference)
    all_ids = [
        c["id"] for c in all_competitors
        if c["id"] != "toolzone_sk" and c["id"] not in own
    ]
    if only:
        competitor_ids = [c for c in all_ids if c in only]
    else:
        competitor_ids = all_ids

    competitor_names = {c["id"]: c.get("name", c["id"]) for c in all_competitors}

    if not output:
        Path("reports").mkdir(exist_ok=True)
        output = Path(f"reports/{manufacturer}_{date.today().isoformat()}.xlsx")

    print(f"Loading {manufacturer!r} data from database…")
    with factory() as session:
        tz_products, matches = _load_data(manufacturer, competitor_ids, min_confidence, session)

    if not tz_products:
        print(f"No ToolZone listings found for manufacturer '{manufacturer}'.")
        return output

    matched_count = sum(1 for p in tz_products if matches.get(p["id"]))
    print(f"  {len(tz_products)} ToolZone products")
    print(f"  {matched_count} with at least one competitor match")
    print(f"  {len(competitor_ids)} competitors: {', '.join(competitor_ids)}")
    print(f"Building workbook…")

    wb = build_workbook(manufacturer, tz_products, matches, competitor_ids, competitor_names)
    wb.save(output)
    print(f"Saved → {output}")
    return output


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Export manufacturer price comparison to a formatted Excel file."
    )
    parser.add_argument(
        "--manufacturer",
        required=True,
        metavar="SLUG",
        help="Manufacturer brand to export (e.g. knipex, wiha)",
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        metavar="PATH",
        help="Output .xlsx path (default: reports/{manufacturer}_YYYY-MM-DD.xlsx)",
    )
    parser.add_argument(
        "--only",
        nargs="+",
        metavar="COMPETITOR_ID",
        help="Include only these competitor IDs (default: all)",
    )
    parser.add_argument(
        "--min-confidence",
        type=float,
        default=0.72,
        metavar="FLOAT",
        help="Minimum match confidence to include (default: 0.72)",
    )
    args = parser.parse_args()

    import logging
    logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(message)s")

    main(
        manufacturer=args.manufacturer,
        output=args.output,
        only=args.only,
        min_confidence=args.min_confidence,
    )
