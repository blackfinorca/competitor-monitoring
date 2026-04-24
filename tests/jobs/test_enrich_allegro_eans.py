from __future__ import annotations

from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from agnaradie_pricing.db.models import Base, Product


def _load_module():
    module_path = Path(__file__).resolve().parents[2] / "jobs" / "enrich_allegro_eans.py"
    spec = spec_from_file_location("enrich_allegro_eans", module_path)
    assert spec is not None
    assert spec.loader is not None
    module = module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _make_products(engine, *products: dict) -> None:
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        for product in products:
            session.add(Product(**product))
        session.commit()


def test_load_source_rows_supports_csv_and_xlsx(tmp_path) -> None:
    enrich = _load_module()

    csv_path = tmp_path / "allegro.csv"
    csv_path.write_text(
        "ean,title,our_price_sk\n8581000000001,Product A,11.5\n8581000000002,Product B,12.5\n",
        encoding="utf-8",
    )

    xlsx_path = tmp_path / "allegro.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["products_ean", "title", "price_sk"])
    ws.append(["8581000000001", "Product A", 11.5])
    ws.append(["8581000000002", "Product B", 12.5])
    wb.save(xlsx_path)
    wb.close()

    csv_rows = enrich.load_source_rows(csv_path)
    xlsx_rows = enrich.load_source_rows(xlsx_path)

    assert [(row["ean"], row["title"], row["allegro_price_sk"]) for row in csv_rows] == [
        ("8581000000001", "Product A", "11.5"),
        ("8581000000002", "Product B", "12.5"),
    ]
    assert [(row["ean"], row["title"], row["allegro_price_sk"]) for row in xlsx_rows] == [
        ("8581000000001", "Product A", "11.5"),
        ("8581000000002", "Product B", "12.5"),
    ]


def test_main_backfills_unique_exact_title_match(tmp_path, monkeypatch) -> None:
    enrich = _load_module()
    engine = create_engine("sqlite+pysqlite:///:memory:")
    _make_products(
        engine,
        {
            "sku": "AG-1",
            "brand": "Bosch",
            "mpn": "GBH18V21",
            "ean": None,
            "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo",
        },
    )

    csv_path = tmp_path / "allegro.csv"
    csv_path.write_text(
        "ean,title,our_price_sk\n4006825599999,Bosch GBH 18V-21 Aku Vrtacie Kladivo,99.0\n",
        encoding="utf-8",
    )
    report_path = tmp_path / "report.csv"

    monkeypatch.setattr(enrich, "Settings", lambda: type("S", (), {"database_url": "sqlite+pysqlite:///:memory:"})())
    monkeypatch.setattr(enrich, "make_engine", lambda settings: engine)

    rc = enrich.main(
        input_path=csv_path,
        dry_run=False,
        limit=None,
        batch_size=100,
        llm=False,
        report_path=report_path,
    )

    assert rc == 0
    with Session(engine) as session:
        product = session.scalar(select(Product).where(Product.sku == "AG-1"))
        assert product is not None
        assert product.ean == "4006825599999"

    report = report_path.read_text(encoding="utf-8")
    assert "matched" in report
    assert "exact_title" in report


def test_main_does_not_overwrite_existing_ean(tmp_path, monkeypatch) -> None:
    enrich = _load_module()
    engine = create_engine("sqlite+pysqlite:///:memory:")
    _make_products(
        engine,
        {
            "sku": "AG-1",
            "brand": "Bosch",
            "mpn": "GBH18V21",
            "ean": "1111111111111",
            "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo",
        },
    )

    csv_path = tmp_path / "allegro.csv"
    csv_path.write_text(
        "ean,title,our_price_sk\n4006825599999,Bosch GBH 18V-21 Professional Aku Kladivo,99.0\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(enrich, "Settings", lambda: type("S", (), {"database_url": "sqlite+pysqlite:///:memory:"})())
    monkeypatch.setattr(enrich, "make_engine", lambda settings: engine)

    enrich.main(
        input_path=csv_path,
        dry_run=False,
        limit=None,
        batch_size=100,
        llm=False,
        report_path=tmp_path / "report.csv",
    )

    with Session(engine) as session:
        product = session.scalar(select(Product).where(Product.sku == "AG-1"))
        assert product is not None
        assert product.ean == "1111111111111"


def test_main_uses_deterministic_vector_thresholds_before_writing(tmp_path, monkeypatch) -> None:
    enrich = _load_module()
    engine = create_engine("sqlite+pysqlite:///:memory:")
    _make_products(
        engine,
        {
            "sku": "AG-1",
            "brand": "Bosch",
            "mpn": "GBH18V21",
            "ean": None,
            "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo",
        },
        {
            "sku": "AG-2",
            "brand": "Bosch",
            "mpn": "GBH18V26",
            "ean": None,
            "title": "Bosch GBH 18V-26 Aku Vrtacie Kladivo",
        },
    )

    class FakeVectorIndex:
        def __init__(self, products, *, embedder=None):
            del embedder
            self.products = products

        def search(self, listing, *, limit=20):
            del listing, limit
            return self.products

        def search_with_scores(self, listing, *, limit=20):
            del listing, limit
            return [
                (self.products[0], 0.95),
                (self.products[1], 0.88),
            ]

    csv_path = tmp_path / "allegro.csv"
    csv_path.write_text(
        "ean,title,our_price_sk\n4006825599999,Bosch GBH 18V-21 Aku Vrtacie Kladivo,99.0\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(enrich, "Settings", lambda: type("S", (), {"database_url": "sqlite+pysqlite:///:memory:"})())
    monkeypatch.setattr(enrich, "make_engine", lambda settings: engine)
    monkeypatch.setattr(enrich, "TitleVectorIndex", FakeVectorIndex)

    enrich.main(
        input_path=csv_path,
        dry_run=False,
        limit=None,
        batch_size=100,
        llm=False,
        report_path=tmp_path / "report.csv",
    )

    with Session(engine) as session:
        winner = session.scalar(select(Product).where(Product.sku == "AG-1"))
        loser = session.scalar(select(Product).where(Product.sku == "AG-2"))
        assert winner is not None and loser is not None
        assert winner.ean == "4006825599999"
        assert loser.ean is None


def test_main_uses_llm_only_for_ambiguous_shortlists(tmp_path, monkeypatch) -> None:
    enrich = _load_module()
    engine = create_engine("sqlite+pysqlite:///:memory:")
    _make_products(
        engine,
        {
            "sku": "AG-1",
            "brand": "Bosch",
            "mpn": "GBH18V21",
            "ean": None,
            "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo",
        },
        {
            "sku": "AG-2",
            "brand": "Bosch",
            "mpn": "GBH18V26",
            "ean": None,
            "title": "Bosch GBH 18V-26 Aku Vrtacie Kladivo",
        },
    )

    class FakeVectorIndex:
        def __init__(self, products, *, embedder=None):
            del embedder
            self.products = products

        def search(self, listing, *, limit=20):
            del listing, limit
            return self.products

        def search_with_scores(self, listing, *, limit=20):
            del listing, limit
            return [
                (self.products[0], 0.84),
                (self.products[1], 0.82),
            ]

    seen_candidates: list[int] = []

    def fake_find_best_llm_match(listing, candidates, *, llm_client):
        del listing, llm_client
        seen_candidates.append(len(candidates))
        return candidates[0], ("llm_fuzzy", 0.91)

    csv_path = tmp_path / "allegro.csv"
    csv_path.write_text(
        "ean,title,our_price_sk\n4006825599999,Bosch GBH 18V-21 Professional Aku Kladivo,99.0\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(enrich, "Settings", lambda: type("S", (), {"database_url": "sqlite+pysqlite:///:memory:"})())
    monkeypatch.setattr(enrich, "make_engine", lambda settings: engine)
    monkeypatch.setattr(enrich, "TitleVectorIndex", FakeVectorIndex)
    monkeypatch.setattr(enrich, "find_best_llm_match", fake_find_best_llm_match)

    enrich.main(
        input_path=csv_path,
        dry_run=False,
        limit=None,
        batch_size=100,
        llm=True,
        report_path=tmp_path / "report.csv",
    )

    with Session(engine) as session:
        winner = session.scalar(select(Product).where(Product.sku == "AG-1"))
        assert winner is not None
        assert winner.ean == "4006825599999"

    assert seen_candidates == [2]


def test_main_rejects_duplicate_ean_mapping_to_multiple_products(tmp_path, monkeypatch) -> None:
    enrich = _load_module()
    engine = create_engine("sqlite+pysqlite:///:memory:")
    _make_products(
        engine,
        {
            "sku": "AG-1",
            "brand": "Bosch",
            "mpn": "GBH18V21",
            "ean": None,
            "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo",
        },
        {
            "sku": "AG-2",
            "brand": "Bosch",
            "mpn": "GBH18V26",
            "ean": None,
            "title": "Bosch GBH 18V-26 Aku Vrtacie Kladivo",
        },
    )

    class FakeVectorIndex:
        def __init__(self, products, *, embedder=None):
            self.products = products

        def search(self, listing, *, limit=20):
            del listing, limit
            return self.products

        def search_with_scores(self, listing, *, limit=20):
            del limit
            title = listing["title"]
            if "21" in title:
                return [(self.products[0], 0.95), (self.products[1], 0.70)]
            return [(self.products[1], 0.95), (self.products[0], 0.70)]

    csv_path = tmp_path / "allegro.csv"
    csv_path.write_text(
        (
            "ean,title,our_price_sk\n"
            "4006825599999,Bosch GBH 18V-21 Aku Vrtacie Kladivo,99.0\n"
            "4006825599999,Bosch GBH 18V-26 Aku Vrtacie Kladivo,109.0\n"
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(enrich, "Settings", lambda: type("S", (), {"database_url": "sqlite+pysqlite:///:memory:"})())
    monkeypatch.setattr(enrich, "make_engine", lambda settings: engine)
    monkeypatch.setattr(enrich, "TitleVectorIndex", FakeVectorIndex)

    enrich.main(
        input_path=csv_path,
        dry_run=False,
        limit=None,
        batch_size=100,
        llm=False,
        report_path=tmp_path / "report.csv",
    )

    with Session(engine) as session:
        products = session.scalars(select(Product).order_by(Product.sku)).all()
        assert [product.ean for product in products] == [None, None]

    report = (tmp_path / "report.csv").read_text(encoding="utf-8")
    assert "duplicate_ean_conflict" in report


def test_main_dry_run_writes_report_without_db_changes(tmp_path, monkeypatch) -> None:
    enrich = _load_module()
    engine = create_engine("sqlite+pysqlite:///:memory:")
    _make_products(
        engine,
        {
            "sku": "AG-1",
            "brand": "Bosch",
            "mpn": "GBH18V21",
            "ean": None,
            "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo",
        },
    )

    csv_path = tmp_path / "allegro.csv"
    csv_path.write_text(
        "ean,title,our_price_sk\n4006825599999,Bosch GBH 18V-21 Aku Vrtacie Kladivo,99.0\n",
        encoding="utf-8",
    )
    report_path = tmp_path / "report.csv"

    monkeypatch.setattr(enrich, "Settings", lambda: type("S", (), {"database_url": "sqlite+pysqlite:///:memory:"})())
    monkeypatch.setattr(enrich, "make_engine", lambda settings: engine)

    enrich.main(
        input_path=csv_path,
        dry_run=True,
        limit=None,
        batch_size=100,
        llm=False,
        report_path=report_path,
    )

    with Session(engine) as session:
        product = session.scalar(select(Product).where(Product.sku == "AG-1"))
        assert product is not None
        assert product.ean is None

    assert report_path.exists()
    assert "matched" in report_path.read_text(encoding="utf-8")


def test_main_flushes_buffer_when_interrupted(tmp_path, monkeypatch) -> None:
    enrich = _load_module()
    engine = create_engine("sqlite+pysqlite:///:memory:")
    _make_products(
        engine,
        {
            "sku": "AG-1",
            "brand": "Bosch",
            "mpn": "GBH18V21",
            "ean": None,
            "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo",
        },
        {
            "sku": "AG-2",
            "brand": "Bosch",
            "mpn": "GBH18V26",
            "ean": None,
            "title": "Bosch GBH 18V-26 Aku Vrtacie Kladivo",
        },
    )

    rows = [
        {"ean": "4006825599999", "title": "Bosch GBH 18V-21 Aku Vrtacie Kladivo", "allegro_price_sk": "99.0"},
        {"ean": "4006825511111", "title": "Bosch GBH 18V-26 Aku Vrtacie Kladivo", "allegro_price_sk": "109.0"},
    ]

    def fake_iter_rows(*args, **kwargs):
        del args, kwargs
        yield rows[0]
        raise KeyboardInterrupt()

    monkeypatch.setattr(enrich, "load_source_rows", lambda path: rows)
    monkeypatch.setattr(enrich, "_iter_source_rows", fake_iter_rows)
    monkeypatch.setattr(enrich, "Settings", lambda: type("S", (), {"database_url": "sqlite+pysqlite:///:memory:"})())
    monkeypatch.setattr(enrich, "make_engine", lambda settings: engine)

    rc = enrich.main(
        input_path=tmp_path / "allegro.csv",
        dry_run=False,
        limit=None,
        batch_size=100,
        llm=False,
        report_path=tmp_path / "report.csv",
    )

    assert rc == 130
    with Session(engine) as session:
        first = session.scalar(select(Product).where(Product.sku == "AG-1"))
        second = session.scalar(select(Product).where(Product.sku == "AG-2"))
        assert first is not None and second is not None
        assert first.ean == "4006825599999"
        assert second.ean is None
